import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from gspread.exceptions import APIError


CANONICAL_COLUMNS = [
    "NO",
    "NO_HULL",
    "TYPE_CAR",
    "DRIVER",
    "FROM",
    "DESTINATION",
    "ACTIVITIES",
    "Date_Departure",
    "Date_Arrival",
    "Distance_Start",
    "Distance_Finish",
    "Distance_Total",
    "BBM,_L",
    "Time_Departure",
    "Time_Arrival",
    "TOTAL_TIME",
]

REQUIRED_TEXT_FIELDS = {
    "NO_HULL": "NO_HULL",
    "TYPE_CAR": "TYPE_CAR",
    "DRIVER": "DRIVER",
    "FROM": "FROM",
    "DESTINATION": "DESTINATION",
    "ACTIVITIES": "ACTIVITIES",
}


def apply_custom_theme() -> None:
    st.markdown(
        """
        <style>
            :root {
                --bg-soft: #0d1524;
                --bg-soft-2: #132037;
                --ink: #e8f0fb;
                --muted: #a9bbd4;
                --card: #16263d;
                --border: #2a3f5e;
            }

            .stApp {
                background: radial-gradient(circle at 25% 8%, #162843 0%, var(--bg-soft-2) 45%, var(--bg-soft) 100%);
                color: var(--ink);
            }

            .block-container {
                background: transparent;
                border-radius: 0;
                padding-top: 1.75rem;
                padding-bottom: 1.2rem;
                overflow: visible;
            }

            [data-testid="stSidebar"] {
                background: linear-gradient(180deg, #0f2747 0%, #15365f 100%);
                border-right: 1px solid rgba(255,255,255,0.08);
            }

            [data-testid="stSidebar"] * {
                color: #e7eef7;
            }

            .app-hero {
                background: linear-gradient(120deg, #0f2747 0%, #1f4d7a 55%, #2a9d8f 100%);
                border-radius: 14px;
                padding: 1rem 1.25rem;
                margin-top: 0.25rem;
                margin-bottom: 1rem;
                color: #f4f8fc;
                box-shadow: 0 10px 30px rgba(15, 39, 71, 0.18);
            }

            .app-hero h1 {
                margin: 0;
                font-size: 1.5rem;
                line-height: 1.2;
                letter-spacing: 0.2px;
            }

            .app-hero p {
                margin: 0.35rem 0 0;
                opacity: 0.92;
                font-size: 0.95rem;
            }

            .meta-card {
                background: var(--card);
                border: 1px solid var(--border);
                border-radius: 12px;
                padding: 0.75rem 0.9rem;
                box-shadow: 0 5px 18px rgba(16, 35, 58, 0.06);
            }

            .meta-title {
                margin: 0;
                color: var(--muted);
                font-size: 0.78rem;
                text-transform: uppercase;
                letter-spacing: 0.8px;
            }

            .meta-value {
                margin: 0.2rem 0 0;
                color: var(--ink);
                font-weight: 650;
                font-size: 1rem;
            }

            .stButton > button {
                border-radius: 10px;
                border: 1px solid var(--border);
                font-weight: 600;
            }

            .stDownloadButton > button {
                border-radius: 10px;
                border: 1px solid #3c587e;
                background: #1b3150;
                color: #dbe8f8;
                font-weight: 600;
            }

            [data-testid="stWidgetLabel"] p {
                color: #dce9f8 !important;
                font-weight: 600;
            }

            .stCaption {
                color: #a8bdd7 !important;
            }

            [data-testid="stNumberInput"] input,
            [data-testid="stTextInput"] input,
            [data-testid="stTextArea"] textarea,
            [data-testid="stDateInput"] input,
            [data-testid="stTimeInput"] input,
            [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
                color: #e8f0fb !important;
            }

            [data-testid="stDataFrame"] {
                border: 1px solid var(--border);
                border-radius: 12px;
                overflow: hidden;
            }

            .section-title {
                color: #dce9f8;
                margin-top: 0.35rem;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_hero(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="app-hero">
            <h1>{title}</h1>
            <p>{subtitle}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_meta_card(label: str, value: str) -> None:
    st.markdown(
        f"""
        <div class="meta-card">
            <p class="meta-title">{label}</p>
            <p class="meta-value">{value}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def normalize_sheet_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize sheet data so app always works with the expected columns."""
    if df is None or df.empty:
        return pd.DataFrame(columns=CANONICAL_COLUMNS)

    work = df.copy()
    work = work.dropna(how="all")

    header_tokens = {
        "NO",
        "NO_HULL",
        "TYPE_CAR",
        "DRIVER",
        "FROM",
        "DESTINATION",
        "ACTIVITIES",
        "DEPARTURE",
        "ARRIVAL",
        "START",
        "FINISH",
        "TOTAL",
    }

    header_row_idx = None
    max_scan = min(len(work), 15)
    for idx in range(max_scan):
        values = {
            str(v).strip().upper()
            for v in work.iloc[idx].tolist()
            if pd.notna(v) and str(v).strip() and str(v).strip().lower() != "none"
        }
        if len(values & header_tokens) >= 6:
            header_row_idx = idx
            break

    if header_row_idx is not None:
        work = work.iloc[header_row_idx + 1 :].reset_index(drop=True)

    work = work.iloc[:, : len(CANONICAL_COLUMNS)].copy()
    while work.shape[1] < len(CANONICAL_COLUMNS):
        work[f"_pad_{work.shape[1]}"] = pd.NA

    work.columns = CANONICAL_COLUMNS
    work = work.dropna(how="all")

    # Kolom NO idealnya integer agar urutan jelas saat tambah data.
    work["NO"] = pd.to_numeric(work["NO"], errors="coerce").astype("Int64")
    return work


def stringify_for_editor(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare dataframe so every editable field is safe for Streamlit editor."""
    editor_df = df.copy()
    for col in editor_df.columns:
        editor_df[col] = editor_df[col].where(editor_df[col].notna(), "")
    return editor_df


def filter_dataframe(df: pd.DataFrame, keyword: str) -> pd.DataFrame:
    """Simple global search across all columns."""
    if not keyword:
        return df
    mask = df.astype(str).apply(lambda s: s.str.contains(keyword, case=False, na=False)).any(axis=1)
    return df[mask]


def normalize_text(value) -> str:
    return str(value).strip().upper() if value is not None else ""


def get_next_no(df: pd.DataFrame) -> int:
    """Get next NO based on the last non-empty NO in current month sheet."""
    numeric_no = pd.to_numeric(df["NO"], errors="coerce")
    valid_no = numeric_no.dropna()
    if valid_no.empty:
        return 1
    return int(valid_no.iloc[-1]) + 1


def get_hull_reference(conn, spreadsheet: str, month_sheets: list[str]):
    """Collect NO_HULL options and default TYPE_CAR map from all months."""
    all_rows = []
    for month_name in month_sheets:
        try:
            month_raw = conn.read(spreadsheet=spreadsheet, worksheet=month_name, ttl=60)
            month_df = normalize_sheet_dataframe(month_raw)
            if month_df.empty:
                continue
            ref_df = month_df[["NO_HULL", "TYPE_CAR"]].copy()
            ref_df = ref_df.dropna(how="all")
            all_rows.append(ref_df)
        except Exception:
            continue

    if not all_rows:
        return [], {}

    merged = pd.concat(all_rows, ignore_index=True)
    merged["NO_HULL"] = merged["NO_HULL"].astype(str).str.strip()
    merged["TYPE_CAR"] = merged["TYPE_CAR"].astype(str).str.strip()
    merged = merged[merged["NO_HULL"] != ""]

    hull_options = sorted(merged["NO_HULL"].drop_duplicates().tolist())

    hull_to_type = {}
    for hull, grp in merged.groupby("NO_HULL"):
        # Buang nilai kosong dan pseudo-null agar pemilihan default tidak error.
        types = []
        for val in grp["TYPE_CAR"].tolist():
            norm = str(val).strip()
            if not norm:
                continue
            if norm.lower() in {"nan", "none", "null", "<na>"}:
                continue
            types.append(norm)

        if not types:
            continue

        # Pakai nilai paling sering sebagai default auto-fill TYPE_CAR.
        counts = pd.Series(types).value_counts()
        if counts.empty:
            continue
        hull_to_type[hull] = str(counts.index[0])

    return hull_options, hull_to_type


def is_rate_limit_error(exc: Exception) -> bool:
    return "429" in str(exc) or "RATE_LIMIT_EXCEEDED" in str(exc)


def get_hull_reference_cached(conn, spreadsheet: str, month_sheets: list[str]):
    """Session cache to avoid hitting Google Sheets read quota repeatedly."""
    cache_key = f"{spreadsheet}|{'|'.join(month_sheets)}"
    cached = st.session_state.get("hull_ref_cache")
    if cached and cached.get("key") == cache_key:
        return cached.get("options", []), cached.get("mapping", {})

    options, mapping = get_hull_reference(conn, spreadsheet, month_sheets)
    st.session_state["hull_ref_cache"] = {
        "key": cache_key,
        "options": options,
        "mapping": mapping,
    }
    return options, mapping


def validate_required_text_values(data: dict) -> list[str]:
    missing = []
    for key, label in REQUIRED_TEXT_FIELDS.items():
        if not str(data.get(key, "")).strip():
            missing.append(label)
    return missing


def get_add_form_defaults() -> dict:
    now_time = datetime.now().time().replace(second=0, microsecond=0)
    return {
        "add_no_hull": "",
        "add_type_car": "",
        "add_driver": "",
        "add_from": "",
        "add_destination": "",
        "add_activities": "",
        "add_date_dep": datetime.today().date(),
        "add_time_dep": now_time,
        "add_date_arr": datetime.today().date(),
        "add_time_arr": now_time,
        "add_dist_start": 0.0,
        "add_dist_finish": 0.0,
    }


def apply_add_form_reset_if_needed() -> None:
    if st.session_state.get("reset_add_form_pending"):
        for k, v in get_add_form_defaults().items():
            st.session_state[k] = v
        st.session_state["type_auto_from_hull"] = ""
        st.session_state["reset_add_form_pending"] = False


def build_styled_excel(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Create a styled Excel file similar to the spreadsheet layout."""
    export_df = df.copy()

    for col in ["Date_Departure", "Date_Arrival"]:
        parsed = pd.to_datetime(export_df[col], errors="coerce")
        export_df[col] = parsed.dt.strftime("%d/%m/%Y")
        export_df[col] = export_df[col].fillna("")

    for col in ["Time_Departure", "Time_Arrival", "TOTAL_TIME"]:
        export_df[col] = export_df[col].fillna("").astype(str)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2, header=False)
        ws = writer.book[sheet_name]

        ws.merge_cells("H1:I1")
        ws.merge_cells("J1:L1")
        ws.merge_cells("N1:P1")

        ws["H1"] = "DATE"
        ws["J1"] = "Distance (KM)"
        ws["N1"] = "Time"

        header_labels = [
            "NO",
            "NO_HULL",
            "TYPE_CAR",
            "DRIVER",
            "FROM",
            "DESTINATION",
            "ACTIVITIES",
            "Departure",
            "Arrival",
            "Start",
            "Finish",
            "Total",
            "BBM,_L",
            "Departure",
            "Arrival",
            "TOTAL_TIME",
        ]
        for col_idx, label in enumerate(header_labels, start=1):
            ws.cell(row=2, column=col_idx, value=label)

        green_fill = PatternFill("solid", fgColor="D9EAD3")
        blue_fill = PatternFill("solid", fgColor="9FC5E8")
        yellow_fill = PatternFill("solid", fgColor="F1C232")

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill_for_col(col_idx: int):
            if col_idx in (8, 9):
                return blue_fill
            if col_idx in (10, 11, 12, 13):
                return yellow_fill
            return green_fill

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

        for row_idx in (1, 2):
            for col_idx in range(1, 17):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_for_col(col_idx)
                cell.font = header_font
                cell.alignment = center
                cell.border = border

        widths = [7, 14, 12, 12, 12, 14, 28, 12, 12, 11, 11, 10, 9, 11, 11, 11]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = border
                if cell.column == 7:
                    cell.alignment = left
                else:
                    cell.alignment = center

        ws.freeze_panes = "A3"

    return buffer.getvalue()

try:
    from streamlit_gsheets import GSheetsConnection
except ImportError:
    GSheetsConnection = None

# --- UI STREAMLIT ---
st.set_page_config(page_title="Rekap Sarana Kendaraan", layout="wide")
apply_custom_theme()
render_hero(
    "Aplikasi Rekap Sarana Kendaraan",
    "Pencatatan operasional kendaraan bulanan dengan alur data terstruktur.",
)

# --- KONEKSI KE GOOGLE SHEETS ---
# Membuat objek koneksi
if GSheetsConnection is None:
    st.error("Dependency st-gsheets-connection belum terpasang. Jalankan install dependency dari requirements.txt.")
    st.stop()

try:
    conn = st.connection("gsheets", type=GSheetsConnection)
except Exception as exc:
    st.error("Koneksi Google Sheets belum siap. Pastikan secret Streamlit sudah diisi dengan benar.")
    st.caption(str(exc))
    st.stop()

gsheets_spreadsheet = ""
try:
    gsheets_spreadsheet = st.secrets.get("connections", {}).get("gsheets", {}).get("spreadsheet", "")
except Exception:
    gsheets_spreadsheet = ""

# Branding sidebar
st.sidebar.image("logo.jpeg", use_container_width=True)

# Fallback aman: spreadsheet bisa diisi dari sidebar jika belum ditetapkan di secrets.
target_spreadsheet = st.sidebar.text_input(
    "Spreadsheet (nama atau URL)",
    value=str(gsheets_spreadsheet).strip(),
    placeholder="Contoh: Rekap Sarana Kendaraan atau https://docs.google.com/spreadsheets/...",
)

if not str(target_spreadsheet).strip():
    st.error("Konfigurasi belum lengkap: isi nilai 'spreadsheet' di sidebar atau Streamlit secrets.")
    st.stop()

# Daftar bulan/sheet yang tersedia (harus sesuai dengan nama tab di Google Sheets)
sheet_names = [
    "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"
]

# Sidebar Menu
st.sidebar.markdown("### Navigasi")
menu = st.sidebar.radio("Pilih Menu", ["Lihat Data", "Tambah Data", "Edit Data", "Hapus Data"])
selected_sheet = st.sidebar.selectbox("Pilih Bulan (Sheet)", sheet_names)

# Tarik data dari Google Sheets dengan TTL lebih longgar agar tidak cepat kena quota.
try:
    df_raw = conn.read(spreadsheet=target_spreadsheet.strip(), worksheet=selected_sheet, ttl=60)
except APIError as exc:
    if is_rate_limit_error(exc):
        st.error("Quota Google Sheets terbatas (429). Tunggu sekitar 1 menit lalu coba lagi.")
        st.info("Tips: kurangi refresh berulang dan gunakan menu seperlunya agar read request tidak cepat habis.")
    else:
        st.error("Gagal membaca Google Sheets.")
    st.caption(str(exc))
    st.stop()

df_current = normalize_sheet_dataframe(df_raw)

meta_col1, meta_col2, meta_col3 = st.columns(3)
with meta_col1:
    render_meta_card("Menu Aktif", menu)
with meta_col2:
    render_meta_card("Bulan Aktif", selected_sheet)
with meta_col3:
    render_meta_card("Jumlah Baris", str(len(df_current)))

if menu == "Lihat Data":
    st.markdown('<h3 class="section-title">Lihat Data</h3>', unsafe_allow_html=True)
    display_df = df_current.rename(
        columns={
            "Date_Departure": "Departure Date",
            "Date_Arrival": "Arrival Date",
            "Distance_Start": "Start (KM)",
            "Distance_Finish": "Finish (KM)",
            "Distance_Total": "Total (KM)",
            "Time_Departure": "Departure Time",
            "Time_Arrival": "Arrival Time",
        }
    )

    search_keyword = st.text_input(
        "Cari data (NO, driver, hull, tujuan, aktivitas, dll)",
        placeholder="Ketik kata kunci pencarian...",
    ).strip()
    filtered_display_df = filter_dataframe(display_df, search_keyword)

    st.subheader(f"Data Rekap - {selected_sheet}")
    st.dataframe(filtered_display_df, use_container_width=True)

    excel_bytes = build_styled_excel(df_current, selected_sheet)
    st.download_button(
        label=f"Download Data {selected_sheet} (Excel)",
        data=excel_bytes,
        file_name=f"Rekap_Kendaraan_{selected_sheet}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if menu == "Tambah Data":
    st.markdown('<h3 class="section-title">Tambah Data</h3>', unsafe_allow_html=True)
    st.subheader(f"Input Data Baru - {selected_sheet}")
    next_no = get_next_no(df_current)
    apply_add_form_reset_if_needed()
    try:
        hull_options, hull_to_type = get_hull_reference_cached(conn, target_spreadsheet.strip(), sheet_names)
    except APIError as exc:
        if is_rate_limit_error(exc):
            st.warning("Referensi NO_HULL lintas bulan sementara tidak bisa di-refresh karena quota read habis.")
            cached = st.session_state.get("hull_ref_cache", {})
            hull_options = cached.get("options", [])
            hull_to_type = cached.get("mapping", {})
        else:
            raise

    if st.button("Refresh Referensi NO_HULL"):
        st.session_state.pop("hull_ref_cache", None)
        st.rerun()

    add_state_defaults = get_add_form_defaults()
    for k, v in add_state_defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    if st.session_state.get("add_sheet_context") != selected_sheet:
        st.session_state["add_sheet_context"] = selected_sheet

    col1, col2, col3 = st.columns(3)

    with col1:
        st.number_input("NO (Auto)", min_value=1, step=1, value=next_no, disabled=True)
        hull_choices = [""] + hull_options
        current_hull = st.session_state.get("add_no_hull", "")
        if current_hull not in hull_choices:
            current_hull = ""
        st.selectbox(
            "NO_HULL",
            options=hull_choices,
            index=hull_choices.index(current_hull),
            key="add_no_hull",
            help="Dropdown diambil dari data bulan-bulan sebelumnya.",
        )

        selected_hull = st.session_state.get("add_no_hull", "").strip()
        if selected_hull:
            default_type = hull_to_type.get(selected_hull, "")
            last_auto_hull = st.session_state.get("type_auto_from_hull", "")
            if default_type and last_auto_hull != selected_hull:
                st.session_state["add_type_car"] = default_type
                st.session_state["type_auto_from_hull"] = selected_hull

        st.text_input("TYPE_CAR", key="add_type_car")
        st.caption("TYPE_CAR otomatis terisi dari NO_HULL, tapi tetap bisa diedit manual.")
        st.text_input("DRIVER", key="add_driver")
        st.text_input("FROM", key="add_from")
        st.text_input("DESTINATION", key="add_destination")
        st.text_input("ACTIVITIES", key="add_activities")

    with col2:
        st.markdown("**DATE & TIME**")
        st.date_input("Date Departure", key="add_date_dep")
        st.time_input("Time Departure", key="add_time_dep")
        st.date_input("Date Arrival", key="add_date_arr")
        st.time_input("Time Arrival", key="add_time_arr")

    with col3:
        st.markdown("**DISTANCE (KM)**")
        st.number_input("Distance Start", min_value=0.0, step=0.1, key="add_dist_start")
        st.number_input("Distance Finish", min_value=0.0, step=0.1, key="add_dist_finish")

    submit_btn = st.button("Simpan Data", type="primary")

    if submit_btn:
        no = get_next_no(df_current)
        no_hull = st.session_state["add_no_hull"].strip()
        type_car = st.session_state["add_type_car"].strip()
        driver = st.session_state["add_driver"].strip()
        from_loc = st.session_state["add_from"].strip()
        destination = st.session_state["add_destination"].strip()
        activities = st.session_state["add_activities"].strip()
        date_dep = st.session_state["add_date_dep"]
        time_dep = st.session_state["add_time_dep"]
        date_arr = st.session_state["add_date_arr"]
        time_arr = st.session_state["add_time_arr"]
        dist_start = float(st.session_state["add_dist_start"])
        dist_finish = float(st.session_state["add_dist_finish"])

        dist_total = dist_finish - dist_start
        bbm_l = dist_total / 10.0

        datetime_dep = datetime.combine(date_dep, time_dep)
        datetime_arr = datetime.combine(date_arr, time_arr)
        time_diff = datetime_arr - datetime_dep

        total_seconds = int(time_diff.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        time_total_str = f"{hours:02d}:{minutes:02d}"

        missing_fields = validate_required_text_values(
            {
                "NO_HULL": no_hull,
                "TYPE_CAR": type_car,
                "DRIVER": driver,
                "FROM": from_loc,
                "DESTINATION": destination,
                "ACTIVITIES": activities,
            }
        )

        if missing_fields:
            st.error(f"Field wajib belum diisi: {', '.join(missing_fields)}")
        elif dist_finish < dist_start:
            st.error("Distance Finish tidak boleh lebih kecil dari Distance Start!")
        elif datetime_arr < datetime_dep:
            st.error("Waktu Arrival tidak boleh lebih awal dari Departure!")
        else:
            new_data = {
                "NO": no,
                "NO_HULL": no_hull,
                "TYPE_CAR": type_car,
                "DRIVER": driver,
                "FROM": from_loc,
                "DESTINATION": destination,
                "ACTIVITIES": activities,
                "Date_Departure": date_dep.strftime("%Y-%m-%d"),
                "Date_Arrival": date_arr.strftime("%Y-%m-%d"),
                "Distance_Start": dist_start,
                "Distance_Finish": dist_finish,
                "Distance_Total": dist_total,
                "BBM,_L": bbm_l,
                "Time_Departure": time_dep.strftime("%H:%M"),
                "Time_Arrival": time_arr.strftime("%H:%M"),
                "TOTAL_TIME": time_total_str,
            }

            new_row_df = pd.DataFrame([new_data])
            updated_df = pd.concat([df_current, new_row_df], ignore_index=True)
            conn.update(spreadsheet=target_spreadsheet.strip(), worksheet=selected_sheet, data=updated_df)

            st.success("Data berhasil di-push ke Google Sheets.")
            st.cache_data.clear()
            st.session_state["reset_add_form_pending"] = True
            st.rerun()

if menu == "Edit Data":
    st.markdown('<h3 class="section-title">Edit Data</h3>', unsafe_allow_html=True)
    st.subheader(f"Pembaruan Data - {selected_sheet}")
    st.caption("Centang satu baris untuk diedit, lalu isi form edit di bawah.")

    pick_df = stringify_for_editor(df_current)
    pick_df.insert(0, "EDIT", False)

    edit_picker_df = st.data_editor(
        pick_df,
        use_container_width=True,
        num_rows="fixed",
        key=f"edit_pick_{selected_sheet}",
        column_config={
            "EDIT": st.column_config.CheckboxColumn("EDIT", help="Centang 1 baris untuk edit", default=False)
        },
        disabled=CANONICAL_COLUMNS,
    )

    selected_edit_idx = edit_picker_df.index[edit_picker_df["EDIT"] == True].tolist()
    if len(selected_edit_idx) > 1:
        st.warning("Pilih satu baris saja untuk diedit.")
    elif len(selected_edit_idx) == 1:
        row_idx = int(selected_edit_idx[0])
        row_data = df_current.loc[row_idx].to_dict()

        edit_state_defaults = {
            "edit_no": int(pd.to_numeric(row_data.get("NO"), errors="coerce") if pd.notna(row_data.get("NO")) else 0),
            "edit_no_hull": str(row_data.get("NO_HULL", "") or ""),
            "edit_type_car": str(row_data.get("TYPE_CAR", "") or ""),
            "edit_driver": str(row_data.get("DRIVER", "") or ""),
            "edit_from": str(row_data.get("FROM", "") or ""),
            "edit_destination": str(row_data.get("DESTINATION", "") or ""),
            "edit_activities": str(row_data.get("ACTIVITIES", "") or ""),
            "edit_date_dep": pd.to_datetime(row_data.get("Date_Departure"), errors="coerce").date() if pd.notna(pd.to_datetime(row_data.get("Date_Departure"), errors="coerce")) else datetime.today().date(),
            "edit_time_dep": pd.to_datetime(str(row_data.get("Time_Departure", "00:00")), format="%H:%M", errors="coerce").time() if pd.notna(pd.to_datetime(str(row_data.get("Time_Departure", "00:00")), format="%H:%M", errors="coerce")) else datetime.now().time().replace(second=0, microsecond=0),
            "edit_date_arr": pd.to_datetime(row_data.get("Date_Arrival"), errors="coerce").date() if pd.notna(pd.to_datetime(row_data.get("Date_Arrival"), errors="coerce")) else datetime.today().date(),
            "edit_time_arr": pd.to_datetime(str(row_data.get("Time_Arrival", "00:00")), format="%H:%M", errors="coerce").time() if pd.notna(pd.to_datetime(str(row_data.get("Time_Arrival", "00:00")), format="%H:%M", errors="coerce")) else datetime.now().time().replace(second=0, microsecond=0),
            "edit_dist_start": float(pd.to_numeric(row_data.get("Distance_Start"), errors="coerce") if pd.notna(pd.to_numeric(row_data.get("Distance_Start"), errors="coerce")) else 0.0),
            "edit_dist_finish": float(pd.to_numeric(row_data.get("Distance_Finish"), errors="coerce") if pd.notna(pd.to_numeric(row_data.get("Distance_Finish"), errors="coerce")) else 0.0),
        }

        state_row_key = f"{selected_sheet}_{row_idx}"
        if st.session_state.get("edit_context") != state_row_key:
            st.session_state["edit_context"] = state_row_key
            for k, v in edit_state_defaults.items():
                st.session_state[k] = v

        st.markdown("### Form Edit")
        c1, c2, c3 = st.columns(3)

        with c1:
            st.number_input("NO", min_value=1, step=1, key="edit_no", disabled=True)
            st.text_input("NO_HULL", key="edit_no_hull")
            st.text_input("TYPE_CAR", key="edit_type_car")
            st.text_input("DRIVER", key="edit_driver")
            st.text_input("FROM", key="edit_from")
            st.text_input("DESTINATION", key="edit_destination")
            st.text_input("ACTIVITIES", key="edit_activities")

        with c2:
            st.markdown("**DATE & TIME**")
            st.date_input("Date Departure", key="edit_date_dep")
            st.time_input("Time Departure", key="edit_time_dep")
            st.date_input("Date Arrival", key="edit_date_arr")
            st.time_input("Time Arrival", key="edit_time_arr")

        with c3:
            st.markdown("**DISTANCE (KM)**")
            st.number_input("Distance Start", min_value=0.0, step=0.1, key="edit_dist_start")
            st.number_input("Distance Finish", min_value=0.0, step=0.1, key="edit_dist_finish")

        if st.button("Simpan Edit Baris", type="primary"):
            edit_payload = {
                "NO": int(st.session_state["edit_no"]),
                "NO_HULL": st.session_state["edit_no_hull"].strip(),
                "TYPE_CAR": st.session_state["edit_type_car"].strip(),
                "DRIVER": st.session_state["edit_driver"].strip(),
                "FROM": st.session_state["edit_from"].strip(),
                "DESTINATION": st.session_state["edit_destination"].strip(),
                "ACTIVITIES": st.session_state["edit_activities"].strip(),
                "Date_Departure": st.session_state["edit_date_dep"],
                "Time_Departure": st.session_state["edit_time_dep"],
                "Date_Arrival": st.session_state["edit_date_arr"],
                "Time_Arrival": st.session_state["edit_time_arr"],
                "Distance_Start": float(st.session_state["edit_dist_start"]),
                "Distance_Finish": float(st.session_state["edit_dist_finish"]),
            }

            missing_fields = validate_required_text_values(edit_payload)
            if missing_fields:
                st.error(f"Field wajib belum diisi: {', '.join(missing_fields)}")
            else:
                dt_dep = datetime.combine(edit_payload["Date_Departure"], edit_payload["Time_Departure"])
                dt_arr = datetime.combine(edit_payload["Date_Arrival"], edit_payload["Time_Arrival"])
                if edit_payload["Distance_Finish"] < edit_payload["Distance_Start"]:
                    st.error("Distance Finish tidak boleh lebih kecil dari Distance Start!")
                elif dt_arr < dt_dep:
                    st.error("Waktu Arrival tidak boleh lebih awal dari Departure!")
                else:
                    dist_total = edit_payload["Distance_Finish"] - edit_payload["Distance_Start"]
                    bbm_l = dist_total / 10.0
                    total_seconds = int((dt_arr - dt_dep).total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes, _ = divmod(remainder, 60)
                    total_time = f"{hours:02d}:{minutes:02d}"

                    df_updated = df_current.copy()
                    df_updated.loc[row_idx, "NO"] = edit_payload["NO"]
                    df_updated.loc[row_idx, "NO_HULL"] = edit_payload["NO_HULL"]
                    df_updated.loc[row_idx, "TYPE_CAR"] = edit_payload["TYPE_CAR"]
                    df_updated.loc[row_idx, "DRIVER"] = edit_payload["DRIVER"]
                    df_updated.loc[row_idx, "FROM"] = edit_payload["FROM"]
                    df_updated.loc[row_idx, "DESTINATION"] = edit_payload["DESTINATION"]
                    df_updated.loc[row_idx, "ACTIVITIES"] = edit_payload["ACTIVITIES"]
                    df_updated.loc[row_idx, "Date_Departure"] = edit_payload["Date_Departure"].strftime("%Y-%m-%d")
                    df_updated.loc[row_idx, "Date_Arrival"] = edit_payload["Date_Arrival"].strftime("%Y-%m-%d")
                    df_updated.loc[row_idx, "Distance_Start"] = edit_payload["Distance_Start"]
                    df_updated.loc[row_idx, "Distance_Finish"] = edit_payload["Distance_Finish"]
                    df_updated.loc[row_idx, "Distance_Total"] = dist_total
                    df_updated.loc[row_idx, "BBM,_L"] = bbm_l
                    df_updated.loc[row_idx, "Time_Departure"] = edit_payload["Time_Departure"].strftime("%H:%M")
                    df_updated.loc[row_idx, "Time_Arrival"] = edit_payload["Time_Arrival"].strftime("%H:%M")
                    df_updated.loc[row_idx, "TOTAL_TIME"] = total_time

                    conn.update(spreadsheet=target_spreadsheet.strip(), worksheet=selected_sheet, data=df_updated)
                    st.success("Perubahan edit berhasil disimpan ke Google Sheets.")
                    st.cache_data.clear()
                    st.rerun()
    else:
        st.info("Pilih satu baris pada checkbox EDIT untuk membuka form edit.")

if menu == "Hapus Data":
    st.markdown('<h3 class="section-title">Hapus Data</h3>', unsafe_allow_html=True)
    st.subheader(f"Penghapusan Data - {selected_sheet}")
    st.caption("Centang baris yang ingin dihapus, lalu klik tombol hapus.")

    delete_df = stringify_for_editor(df_current)
    delete_df.insert(0, "HAPUS", False)

    delete_editor_df = st.data_editor(
        delete_df,
        use_container_width=True,
        num_rows="fixed",
        key=f"delete_only_{selected_sheet}",
        column_config={
            "HAPUS": st.column_config.CheckboxColumn("HAPUS", help="Centang untuk hapus baris", default=False)
        },
        disabled=CANONICAL_COLUMNS,
    )

    if st.button("Hapus Baris Terpilih", type="primary"):
        rows_to_keep = delete_editor_df[delete_editor_df["HAPUS"] != True].copy()
        rows_to_keep = rows_to_keep.drop(columns=["HAPUS"])

        for col in CANONICAL_COLUMNS:
            if col not in rows_to_keep.columns:
                rows_to_keep[col] = ""
        rows_to_keep = rows_to_keep[CANONICAL_COLUMNS]

        rows_to_keep["NO"] = pd.to_numeric(rows_to_keep["NO"], errors="coerce").astype("Int64")
        for col in ["Distance_Start", "Distance_Finish", "Distance_Total", "BBM,_L"]:
            rows_to_keep[col] = pd.to_numeric(rows_to_keep[col], errors="coerce")

        conn.update(spreadsheet=target_spreadsheet.strip(), worksheet=selected_sheet, data=rows_to_keep)
        st.success("Data terpilih berhasil dihapus dari Google Sheets.")
        st.cache_data.clear()
        st.rerun()